import openpyxl, os, pprint
from openpyxl.styles import Font
bold = Font(bold=True)
import productsData         # products => their prices and calories
import menusData            # menus => their recipes => their products => products quantity for

# how the shopping list data structure looks like
"""
shoppingListData = {'menu1': {'product1': 0, 'product2': 0},
                       'menu2': {'product2': 0, 'product3': 0}}
"""

shoppingListData = {}

for menu in os.listdir():
    if menu.startswith('меню'):

        shoppingListData.setdefault(menu, {})

        menuShoppingList = openpyxl.Workbook()
        menuShoppingListSheet = menuShoppingList.active

        menuShoppingListSheet['A1'] = 'назва меню'
        menuShoppingListSheet['B1'] = 'заг кть'
        menuShoppingListSheet['C1'] = 'заг варт'
        menuShoppingListSheet['A3'] = 'продукт'
        menuShoppingListSheet['B3'] = 'к-ть'
        menuShoppingListSheet['C3'] = 'вартість'
        menuShoppingListSheet['D3'] = 'ціна'
        menuShoppingListSheet['E3'] = '% к-ті'
        menuShoppingListSheet['F3'] = '% вартості'
        menuShoppingListSheet['G3'] = '% ціни'
        menuShoppingListSheet['A2'] = menu[:-5]

        menuShoppingListSheet['A1'].font = bold
        menuShoppingListSheet['B1'].font = bold
        menuShoppingListSheet['C1'].font = bold
        menuShoppingListSheet['A3'].font = bold
        menuShoppingListSheet['B3'].font = bold
        menuShoppingListSheet['C3'].font = bold
        menuShoppingListSheet['D3'].font = bold
        menuShoppingListSheet['E3'].font = bold
        menuShoppingListSheet['F3'].font = bold
        menuShoppingListSheet['G3'].font = bold

        menuWb = openpyxl.load_workbook(menu)
        menuWbSheet = menuWb.active

        for menuRecipes in range(5, menuWbSheet.max_row + 1):
            if menuWbSheet.cell(row=menuRecipes, column=1).value in menusData.menusData[menu]:   # if menu recipe name in menuData.menu keys
                for product in menusData.menusData[menu][menuWbSheet.cell(row=menuRecipes, column=1).value]: # for product name in menuDate.menu.recipeName
                    shoppingListData[menu].setdefault(product, 0)
                    shoppingListData[menu][product] += menusData.menusData[menu][menuWbSheet.cell(row=menuRecipes, column=1).value][product]

        totalShoppingListWeight = 0
        totalShoppingListCost = 0
        totalShoppingListPrice = 0

        for products in shoppingListData:
            # product name in shopping list
            menuShoppingListSheet.cell(row=menuShoppingListSheet.max_row + 1, column=1).value = products

            # product weight in shopping list
            menuShoppingListSheet.cell(row=menuShoppingListSheet.max_row, column=2).value = shoppingListData[menu][products]
            totalShoppingListWeight += menuShoppingListSheet.cell(row=menuShoppingListSheet.max_row, column=2).value

            # product price in shopping list
            menuShoppingListSheet.cell(row=menuShoppingListSheet.max_row, column=4).value = productsData.productsData[products]['price']
            totalShoppingListPrice += menuShoppingListSheet.cell(row=menuShoppingListSheet.max_row, column=4).value

            # product cost in shopping list
            menuShoppingListSheet.cell(row=menuShoppingListSheet.max_row, column=3).value = menuShoppingListSheet.cell(row=menuShoppingListSheet.max_row, column=2).value * productsData.productsData[products]['price'] / 1000
            totalShoppingListCost += menuShoppingListSheet.cell(row=menuShoppingListSheet.max_row, column=3).value


        menuShoppingListSheet['B2'] = totalShoppingListWeight
        menuShoppingListSheet['C2'] = totalShoppingListCost
        menuShoppingListSheet['D2'] = totalShoppingListPrice

        menuShoppingList.save('тест шопінг  ' + menu)

print(pprint.pformat(shoppingListData))
